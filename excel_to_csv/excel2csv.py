#! usr/bin/python3

# excel2csv.py - Converts all the excel files into csv
# files for a given directory

# USAGE: python3 excel2csv.py <directory> - Searches <directory>
# for all .xlsx files and takes the basic cell inputs and
# transfers them to a .csv file for each sheet in the
# Workbook into the same directory.


import csv, openpyxl, os, sys
from openpyxl.utils import get_column_letter


def get_directory(*arg):

    """This function reads in the command line argument
    to get the directory containing the .xlsx files to
    convert into .csv files. If an improper directory
    address, which includes no address specified, is
    given, then the script prompts the user to input a
    directory address until a working one is provided."""

    valid_directory = False
    # If no command line argument for directory is provided
    # prompt user for one.
    if len(arg) < 2:
        print('No directory address provided in command line. Please refer to usage notes:')
        print('USAGE: python3 excel2csv.py <directory> - '
              'Searches <directory> for all .xlsx\nfiles and '
              'takes the basic cell inputs and transfers them '
              'to a .csv file for\neach sheet in the Workbook '
              'into the same directory.')
        directory_address = 'Not Specified'

    else:
        directory_address = ' '.join(sys.argv[1:])

    # Prompt user to enter a directory path until a
    # valid path is entered.
    while not valid_directory:
        if os.path.exists(directory_address):
            abs_path = os.path.abspath(directory_address)
            print('The directory path: %s is valid.\n' % abs_path)
            valid_directory = True

        else:
            print('The directory path: %s is invalid.\n' % directory_address)
            directory_address = input('Enter the directory containing .xlsx files:\n')

    return abs_path


path_name = get_directory(sys.argv)
csv_folder_name = os.path.join((os.path.join(os.getenv('HOME'), 'Desktop')), 'xlsx2csv')
os.makedirs(csv_folder_name, exist_ok=True)

for excelFile in os.listdir(path_name):
    # Skip non-xlsx files, load the workbook object.
    if not excelFile.endswith('.xlsx'):
        continue

    wb = openpyxl.load_workbook(os.path.join(path_name, excelFile))
    print('Preparing to write the file: %s' % excelFile)

    for sheetName in wb.sheetnames:
        # Loop through every sheet in the workbook.
        sheet = wb[sheetName]
        print('Preparing sheet: %s for %s' % (sheetName, excelFile))
        # Create the CSV filename from the Excel filename and sheet title.
        excelFileName = excelFile[:len(excelFile) - 6]
        csvFileName = excelFileName + '_' + sheetName + '.csv'
        csvFilePath = os.path.join(csv_folder_name, csvFileName)
        csvFile = open(csvFilePath, 'w', newline='')
        print('Writing to: %s\n' % csvFilePath)
        # Create the csv.writer object for this CSV file.
        csvWriter = csv.writer(csvFile)

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []    # append each cell to this list

            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                rowData.append(sheet[get_column_letter(colNum) + str(rowNum)].value)
            # Write the rowData list to the CSV file.
            csvWriter.writerow(rowData)

        print('The cell data has been written to: %s' % csvFilePath)
        csvFile.close()

print('The cell data from the .xlsx files in: %s have been written to: %s' % (path_name, csv_folder_name))
