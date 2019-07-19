#! usr/bin/python3

# multiplicationtable.py - This script takes a number N and
# creates a N x N multiplication table in Excel.

# USAGE: python3 multiplicationtable.py <N> - Takes N from the
# command line argument and creates a N x N multiplication
# table in an Excel spreadsheet.


import openpyxl, os, sys
from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl.styles import Font


def get_n(*arg):

    """This function reads the command line argument to check
    for the N value to create the N x N table. The N value must
    be a non-negative integer. If an invalid value is inputted,
    the script will prompt the user to re-enter N values until
    a valid one is inputted.
    """

    # Prompt user for non-negative integer N value if current
    # value is invalid.
    valid_n = False
    n_value = arg
    while not valid_n:
        # Check if N value is valid
        try:
            if isinstance(n_value, int) and arg >= 0:
                valid_n = True
                print('The N value: %s' % n_value)
            else:
                valid_n = False
                n_value = int(input('Enter a non-negative integer value for N: '))
        except Exception as err:
            print('Invalid N value.\nAn error has occurred: %s' % err)
            n_value = int(input('Enter a non-negative integer value for N: '))

    return n_value


def create_excel_table(n_value):

    """This function creates a N x N multiplication table in an Excel
    spreadsheet, which is saved as: multiplication_table.xlsx
    in the Desktop folder.
    """

    # Create the multiplication table in Excel
    wb = openpyxl.Workbook()
    table_sheet = wb.get_sheet_names[1]
    table_sheet.title = 'Multiplication Table'

    # Populate the multiplication table for N x N values
    for column_value in range(n_value):
        column_index = get_column_letter(column_value + 2)
        column_multiplication_factor = column_value + 1
        for row_value in range(n_value):
            row_index = str(row_value + 2)
            row_multiplication_factor = row_value + 1
            table_sheet[column_index + row_index] = column_multiplication_factor*row_multiplication_factor

    # List the multiplication factors
    for index in range(n_value):
        table_sheet['A' + str(index + 2)] = index + 1
        table_sheet['A' + str(index + 2)].font = Font(bold=True)
        table_sheet[get_column_letter(index + 2) + '1'] = index + 1
        table_sheet[get_column_letter(index + 2) + '1'].font = Font(bold=True)

    # Apply freeze panes to the multiplication factors
    table_sheet.freeze_panes = 'B2'

    # Save the Excel spreadsheet in Desktop
    folder_path = os.path.join(os.path.join(os.getenv('HOME'), 'Desktop'), 'multiplication_table.xlsx')
    wb.save(folder_path)
    sys.exit('File saved to: %s' % folder_path)


number = get_n(sys.argv[1])
create_excel_table(number)
